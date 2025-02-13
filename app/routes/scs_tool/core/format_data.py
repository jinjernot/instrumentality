from openpyxl.styles import PatternFill,Font

from config import SCS_REGULAR_FILE_PATH, SCS_GRANULAR_FILE_PATH

import openpyxl

def format_data():
    """This function is used to format the data in the Excel file, bold headers, adjust the column width, and highlight errors."""

    # Load Workbook and get the active sheet
    wb = openpyxl.load_workbook(SCS_REGULAR_FILE_PATH)  # Server
    worksheet = wb.active

    # Bold and color the headers
    header_fill = PatternFill(start_color='0072C6', end_color='0072C6', fill_type='solid')
    for cell in worksheet[1]:
        cell.fill = header_fill

    # Find the column index for the header "Accuracy"
    accuracy_column = None
    for idx, cell in enumerate(worksheet[1], 1):
        if cell.value == "Accuracy":
            accuracy_column = idx
            break

    if accuracy_column is not None:
        # Loop over all the cells in the "Accuracy" column.
        for cell in worksheet.iter_cols(min_col=accuracy_column, max_col=accuracy_column, min_row=2):
            for c in cell:
                if 'ERROR' in str(c.value):
                    font = c.font
                    c.font = Font(color='FF0000', name=font.name, size=font.size)

    # Save the workbook
    wb.save(SCS_REGULAR_FILE_PATH)  # Server
    
def format_data_granular():
    """This function is used to format the data in the Excel file, bold headers, adjust the column width, and highlight errors."""

    # Load Workbook and get the active sheet
    wb = openpyxl.load_workbook(SCS_GRANULAR_FILE_PATH)  # Server
    worksheet = wb.active

    # Bold and color the headers
    header_fill = PatternFill(start_color='0072C6', end_color='0072C6', fill_type='solid')
    for cell in worksheet[1]:
        cell.fill = header_fill

    # Find the column index for the header "Accuracy"
    accuracy_column = None
    for idx, cell in enumerate(worksheet[1], 1):
        if cell.value == "Accuracy":
            accuracy_column = idx
            break

    if accuracy_column is not None:
        # Loop over all the cells in the "Accuracy" column.
        for cell in worksheet.iter_cols(min_col=accuracy_column, max_col=accuracy_column, min_row=2):
            for c in cell:
                if 'ERROR' in str(c.value):
                    font = c.font
                    c.font = Font(color='FF0000', name=font.name, size=font.size)

    # Save the workbook to a file`.
    wb.save(SCS_GRANULAR_FILE_PATH)  # Server